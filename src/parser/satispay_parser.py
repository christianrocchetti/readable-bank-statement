import openpyxl
from datetime import datetime
from typing import List
from models import Transaction, BackType
from parser.parser import TransactionParser

class SatispayParser(TransactionParser):
    
    def get_source_prefix(self) -> str:
        return "(Satispay)"
    
    def parse_file(self, filepath: str) -> List[Transaction]:
        transactions = []
        
        try:
            workbook = openpyxl.load_workbook(filepath)
            
            if "Transactions" not in workbook.sheetnames:
                print(f"Sheet 'Transactions' not found in file {filepath}. Skipping Satispay transactions.")
                return []
            
            worksheet = workbook["Transactions"]
            
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            rows = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)
            
            for row in rows:
                if self.should_include_transaction(row):
                    transaction = self._create_transaction_from_row(row)
                    transactions.append(transaction)
            
            transactions.sort(key=lambda x: x.date)
            
        except FileNotFoundError:
            print(f"Satispay file {filepath} not found. Skipping Satispay transactions.")
            return []
        except Exception as e:
            print(f"Error parsing Satispay file: {e}")
            return []
            
        return transactions
    
    def should_include_transaction(self, raw_data: dict) -> bool:
        try:
            amount = raw_data.get("Amount", 0.00)
            return amount < 0
            
        except (ValueError, KeyError, AttributeError, TypeError):
            return False
    
    def _create_transaction_from_row(self, row: dict) -> Transaction:
        description = row.get("Name", "").strip()
        
        amount_cell = row.get("Amount", 0.00)
        amount = abs(float(amount_cell))
        
        date_str = row.get("Date", "")
        if isinstance(date_str, str):
            date_only = date_str.split(" ")[0]
            date = datetime.strptime(date_only, "%d/%m/%Y")
        else:
            date = date_str if isinstance(date_str, datetime) else datetime.now()
        
        return Transaction(
            description=description,
            amount=amount,
            date=date,
            source_type=BackType.SATISPAY,
            source_prefix=self.get_source_prefix(),
            raw_data=row
        )